from openpyxl import load_workbook

# my_file = load_workbook("C:\Users\Dell\OneDrive\Desktop\emp1.xlsx")
my_file = load_workbook("C:\\Users\\Dell\\OneDrive\\Desktop\\python programming\\emp1.xlsx")
# my_file = load_workbook(r"C:\Users\Dell\OneDrive\Desktop\emp1.xlsx")


s = my_file["Sheet1"]
max_row = s.max_row
max_col = s.max_column

for i in range(1, max_row + 1):
    # Initialize all values
    val1 = val2 = val3 = val4 = val5 = val6 = val7 = val8 = None

    for col in range(1, max_col + 1):
        Valu = s.cell(row=i, column=col).value

        if col == 1:
            val1 = Valu
        elif col == 2:
            val2 = Valu
        elif col == 3:
            val3 = Valu
        elif col == 4:
            val4 = Valu
        elif col == 5:
            val5 = Valu
        elif col == 6:
            val6 = Valu
        elif col == 7:
            val7 = Valu
        elif col == 8:
            val8 = Valu
        # ignore others

    print(i, val1, val2, val3, val4, val5, val6, val7, val8)
    print("done...")
