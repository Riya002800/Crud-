# utils.py

from openpyxl import load_workbook
from typing import List
from .schemas import UserCreate # type: ignore

def parse_excel(file_path: str) -> List[UserCreate]:
    workbook = load_workbook(my_file="C:\\Users\\Dell\\OneDrive\\Desktop\\python programming\\emp1.xlsx")
    sheet = workbook.active  # This gets the first sheet automatically

    users = []
    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):  # Skip header
        name, email, phone = row[:3]  # Only take the first 3 columns
        if not name or not email:
            continue  # Skip rows with missing data
        user = UserCreate(
            name=name.strip(),
            email=email.strip(),
            phone=phone.strip() if phone else None
        )
        users.append(user)

    return users
